"""Excel 学习计划解析（.xlsx / .xls）。

表头映射（中/英均可）：
- 日期/date、科目/subject、内容/content、开始时间/start_time、结束时间/end_time、
  状态/status、优先级/priority、计划名称/plan_name（首行或首个单元格）
无表头时按位置兜底：列1 日期、列2 科目、列3 内容、列4 开始、列5 结束、列6 优先级。

仅做结构化抽取（非正则解析）；返回行数据由上层归一化/落库。
"""
import datetime

_HEADER_MAP = {
    '日期': 'date', 'date': 'date', '时间': 'date',
    '科目': 'subject', 'subject': 'subject', '学科': 'subject',
    '内容': 'content', 'content': 'content', '任务': 'content', 'task': 'content',
    '开始时间': 'start_time', 'start_time': 'start_time', '开始': 'start_time', 'start': 'start_time',
    '结束时间': 'end_time', 'end_time': 'end_time', '结束': 'end_time', 'end': 'end_time',
    '状态': 'status', 'status': 'status',
    '优先级': 'priority', 'priority': 'priority', '重要程度': 'priority',
}


def _norm_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m月%d日', '%Y%m%d'):
        try:
            return datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s or None


def _norm_time(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.time)):
        return v.strftime('%H:%M')
    s = str(v).strip()
    try:
        return datetime.datetime.strptime(s, '%H:%M').strftime('%H:%M')
    except ValueError:
        try:
            return datetime.datetime.strptime(s, '%H:%M:%S').strftime('%H:%M')
        except ValueError:
            return s or None


def extract_excel_rows(file_storage):
    """从 .xlsx/.xls 提取计划行列表 [{date,subject,content,start_time,end_time,priority}]。"""
    from openpyxl import load_workbook
    file_storage.seek(0)
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    header: list[str | None] = []

    for ridx, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [c for c in row] if row else []
        if not cells or all(c is None or str(c).strip() == '' for c in cells):
            continue
        # 第一行尝试当表头：至少包含一个中英表头词
        if ridx == 0:
            keys = [_HEADER_MAP.get(str(c).strip().lower() if isinstance(c, str) else '', None) for c in cells]
            if any(keys):
                header = keys
                continue
        if not header:
            header = ['date', 'subject', 'content', 'start_time', 'end_time', 'priority']
        # 表头长度不够时按位置兜底
        item = {}
        for i, val in enumerate(cells):
            key = header[i] if i < len(header) else None
            if not key:
                break
            item[key] = val
        # 归一化
        norm = {
            'date': _norm_date(item.get('date')),
            'subject': (str(item.get('subject') or '').strip()) or None,
            'content': (str(item.get('content') or '').strip()) or None,
            'start_time': _norm_time(item.get('start_time')),
            'end_time': _norm_time(item.get('end_time')),
            'status': (str(item.get('status') or '').strip().lower()) or 'pending',
            'priority': (str(item.get('priority') or '').strip().lower()) or None,
        }
        # 跳过完全空行
        if not (norm['date'] or norm['content']):
            continue
        rows.append(norm)
    wb.close()
    return rows


def extract_excel_plan_name(file_storage) -> str | None:
    """尝试从 Excel 第一个单元格/第一个非空文本单元格取计划名称（无则 None）。"""
    from openpyxl import load_workbook
    file_storage.seek(0)
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    name = None
    for row in ws.iter_rows(max_row=5, values_only=True):
        for c in row:
            if isinstance(c, str) and c.strip() and '计划' in c:
                name = c.strip()
                break
        if name:
            break
    wb.close()
    return name
